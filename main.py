import sys

import openpyxl
from PyQt5.QtCore import Qt
from PyQt5 import QtCore, QtGui, QtWidgets

from gui import Ui_MainWindow

from openpyxl.utils import get_column_letter

from emerge import MergeManager, NULL_STR, WARN_STR


class MainWindow(QtWidgets.QMainWindow, Ui_MainWindow):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)

        self.setupUi(self)

        self.mm = MergeManager()

        self.load1.clicked.connect(self.setFile1)
        self.load2.clicked.connect(self.setFile2)

        self.cross1.clicked.connect(self.clearFile1)
        self.cross2.clicked.connect(self.clearFile2)

        self.exact.clicked.connect(lambda: self.setFuzzy(False))
        self.approx.clicked.connect(lambda: self.setFuzzy(True))

        self.row11.textChanged.connect(self.row11changed)
        self.row12.textChanged.connect(self.row12changed)
        self.row21.textChanged.connect(self.row21changed)
        self.row22.textChanged.connect(self.row22changed)
        # self.col11.textChanged.connect(self.col11changed)
        # self.col12.textChanged.connect(self.col12changed)
        # self.col21.textChanged.connect(self.col21changed)
        # self.col22.textChanged.connect(self.col22changed)
        self.columnList1.itemSelectionChanged.connect(self.column1SelectionChanged)
        self.columnList2.itemSelectionChanged.connect(self.column2SelectionChanged)
        
        self.table1.setEditTriggers(QtWidgets.QAbstractItemView.NoEditTriggers)
        self.table2.setEditTriggers(QtWidgets.QAbstractItemView.NoEditTriggers)
        self.table3.setEditTriggers(QtWidgets.QAbstractItemView.NoEditTriggers)

        self.mergeon1.currentIndexChanged.connect(self.mergon1changed)
        self.mergeon2.currentIndexChanged.connect(self.mergon2changed)

        self.algo.currentIndexChanged.connect(self.setAlgo)
        self.thresh.valueChanged.connect(self.setThresh)

        self.outerjoin.clicked.connect(lambda: self.setOutputType("outerjoin"))
        self.leftjoin.clicked.connect(lambda: self.setOutputType("leftjoin"))
        self.innerjoin.clicked.connect(lambda: self.setOutputType("innerjoin"))

        self.merge.clicked.connect(self.mergeAndSave)

        self.showMaximized()
        self.statusBar()
        self.setStatusTip("Ready.")

    #################################################################
    ##  PUBLIC
    #################################################################

    def getFile(self):
        return QtWidgets.QFileDialog.getOpenFileName(self, "Choose Excel File", "", "Excel File (*.xlsx)")[0]

    def setFuzzy(self, fuzzy):
        self.mm.setFuzzy(fuzzy)
        self.algo.setEnabled(fuzzy)
        self.thresh.setEnabled(fuzzy)
        self.updateMerge()

    def setOutputType(self, otype):
        self.mm.setOutputType(otype)
        self.updateMerge()

    def setAlgo(self, algo):
        self.mm.setAlgo(algo)
        self.updateMerge()

    def setThresh(self, otype):
        self.mm.setThresh(otype)
        self.updateMerge()

    #################################################################
    ##  SLOTS
    #################################################################

    def setFile1(self):
        f1 = self.getFile()
        self.load1.setText(f1)
        self.mm.setFile1(f1)
        self.updateColumnList1()  # 更新列选择列表
        self.updateTable1()
        self.row11.setText(str(self.mm.file1.startRow))
        self.row12.setText(str(self.mm.file1.endRow))

    def setFile2(self):
        f2 = self.getFile()
        self.load2.setText(f2)
        self.mm.setFile2(f2)
        self.updateColumnList2()  # 更新列选择列表
        self.updateTable2()
        self.row21.setText(str(self.mm.file2.startRow))
        self.row22.setText(str(self.mm.file2.endRow))

    def clearFile1(self):
        self.load1.setText("Load Main File")
        self.mm.setFile1("")
        self.row11.clear()
        self.row12.clear()
        self.columnList1.clear()
        self.updateTable1()

    def clearFile2(self):
        self.load2.setText("Load Secondary File")
        self.mm.setFile2("")
        self.row21.clear()
        self.row22.clear()
        self.columnList2.clear()
        self.updateTable2()

    def row11changed(self, r):
        self.mm.file1.setStartRow(r)
        self.updateTable1()

    def row12changed(self, r):
        self.mm.file1.setEndRow(r)
        self.updateTable1()

    def row21changed(self, r):
        self.mm.file2.setStartRow(r)
        self.updateTable2()

    def row22changed(self, r):
        self.mm.file2.setEndRow(r)
        self.updateTable2()

    def updateColumnList1(self):
        """更新文件1的列选择列表"""
        self.columnList1.clear()
        if not self.mm.file1.fname:
            return

        all_columns = self.mm.file1.getAllColumns()
        for col_info in all_columns:
            item = QtWidgets.QListWidgetItem(f"{col_info['letter']}: {col_info['header']}")
            item.setData(Qt.UserRole, col_info['index'])
            self.columnList1.addItem(item)
        # 新增：更新文件2列选择列表
        for i in range(self.columnList1.count()):
            self.columnList1.item(i).setSelected(True)

    def updateColumnList2(self):
        """更新文件2的列选择列表"""
        self.columnList2.clear()
        if not self.mm.file2.fname:
            return

        all_columns = self.mm.file2.getAllColumns()
        for col_info in all_columns:
            item = QtWidgets.QListWidgetItem(f"{col_info['letter']}: {col_info['header']}")
            item.setData(Qt.UserRole, col_info['index'])
            self.columnList2.addItem(item)

        for i in range(self.columnList2.count()):
            self.columnList2.item(i).setSelected(True)

    def updateMergeColumns1(self):
        """更新文件1的合并列选择"""
        self.mergeon1.clear()
        if not self.mm.file1.selectedColumns:
            return

        for col_index in self.mm.file1.selectedColumns:
            col_letter = f"列{col_index+1}"
            self.mergeon1.addItem(col_letter)
        # 新增：更新文件2合并列选择

        if self.mm.file1.selectedColumns:
            self.mm.file1.setMergeon(self.mm.file1.selectedColumns[0] + 1)
            self.mergeon1.setCurrentIndex(0)

    def updateMergeColumns2(self):
        """更新文件2的合并列选择"""
        self.mergeon2.clear()
        if not self.mm.file2.selectedColumns:
            return

        for col_index in self.mm.file2.selectedColumns:
            col_letter = f"列{col_index+1}"
            self.mergeon2.addItem(col_letter)

        if self.mm.file2.selectedColumns:
            self.mm.file2.setMergeon(self.mm.file2.selectedColumns[0] + 1)
            self.mergeon2.setCurrentIndex(0)

    # def col11changed(self, r):
    #     self.mm.file1.setStartCol(r)
    #     self.updateTable1()
    #     if self.mm.file1.startCol and self.mm.file1.endCol:
    #         colrange = range(self.mm.file1.startCol, self.mm.file1.endCol + 1)
    #         headers = [get_column_letter(x) for x in colrange]
    #         self.mergeon1.clear()
    #         for col in headers:
    #             self.mergeon1.addItem(col)
    #
    # def col12changed(self, r):
    #     self.mm.file1.setEndCol(r)
    #     self.updateTable1()
    #     if self.mm.file1.startCol and self.mm.file1.endCol:
    #         colrange = range(self.mm.file1.startCol, self.mm.file1.endCol + 1)
    #         headers = [get_column_letter(x) for x in colrange]
    #         self.mergeon1.clear()
    #         for col in headers:
    #             self.mergeon1.addItem(col)
    #
    # def col21changed(self, r):
    #     self.mm.file2.setStartCol(r)
    #     self.updateTable2()
    #     if self.mm.file2.startCol and self.mm.file2.endCol:
    #         colrange = range(self.mm.file2.startCol, self.mm.file2.endCol + 1)
    #         headers = [get_column_letter(x) for x in colrange]
    #         self.mergeon2.clear()
    #         for col in headers:
    #             self.mergeon2.addItem(col)
    #
    # def col22changed(self, r):
    #     self.mm.file2.setEndCol(r)
    #     self.updateTable2()
    #     if self.mm.file2.startCol and self.mm.file2.endCol:
    #         colrange = range(self.mm.file2.startCol, self.mm.file2.endCol + 1)
    #         headers = [get_column_letter(x) for x in colrange]
    #         self.mergeon2.clear()
    #         for col in headers:
    #             self.mergeon2.addItem(col)

    # 新增：文件1列选择改变
    def column1SelectionChanged(self):
        """文件1列选择改变"""
        selected_items = self.columnList1.selectedItems()
        selected_indices = [item.data(Qt.UserRole) for item in selected_items]
        self.mm.file1.setSelectedColumns(selected_indices)
        self.updateTable1()
        self.updateMergeColumns1()
        # 新增：文件2列选择改变

    def column2SelectionChanged(self):
        """文件2列选择改变"""
        selected_items = self.columnList2.selectedItems()
        selected_indices = [item.data(Qt.UserRole) for item in selected_items]
        self.mm.file2.setSelectedColumns(selected_indices)
        self.updateTable2()
        self.updateMergeColumns2()

    def mergon1changed(self, mergeon):
        if mergeon >= 0 and mergeon < len(self.mm.file1.selectedColumns):
            selected_col_index = self.mm.file1.selectedColumns[mergeon]
            self.mm.file1.setMergeon(selected_col_index+1)  # 转换为1-based索引
        self.updateTable1()

    def mergon2changed(self, mergeon):
        if mergeon >= 0 and mergeon < len(self.mm.file2.selectedColumns):
            selected_col_index = self.mm.file2.selectedColumns[mergeon]
            self.mm.file2.setMergeon(selected_col_index+1)  # 转换为1-based索引
        self.updateTable2()

    def mergeAndSave(self):
        fname = QtWidgets.QFileDialog.getSaveFileName(self, "Choose Excel File", "", "Excel File (*.xlsx)")[0]
        self.updateMerge()
        self.mm.save(fname)
        self.setStatusTip("Saved output!")

    #################################################################
    ##  PRIVATE
    #################################################################

    def updateTable1(self):
        self.updateTable(self.table1, self.mm.file1, 1)
        self.setStatusTip("Updated File 1")

    def updateTable2(self):
        self.updateTable(self.table2, self.mm.file2, 2)
        self.setStatusTip("Updated File 2")

    def updateTable(self, table, file, num):
        n = len(file.tableData)

        m = len(file.tableData[0]) if n else 0
        table.setRowCount(n)
        table.setColumnCount(m)

        # 设置列标题
        if file.startRow and file.endRow:
            headers = [str(x) for x in range(file.startRow, file.endRow + 1)]
        else:
            headers = [str(x) for x in range(1,n+1)] if n>0 else []
        table.setVerticalHeaderLabels(headers)

        # 设置水平列标题
        if hasattr(file, 'getColumnHeaders'):
            headers = file.getColumnHeaders()
        else:
            headers = [f"列{i+1}" for i in range(m)] if m>0 else []
        table.setHorizontalHeaderLabels(headers)

        for i in range(n):
            for j in range(m):
                data = file.tableData[i][j]
                item = QtWidgets.QTableWidgetItem(data)
                item.setFlags(Qt.ItemIsEnabled)
                if data == NULL_STR:
                    item.setForeground(QtGui.QBrush(QtGui.QColor('red')))
                if data.startswith(WARN_STR):
                    item.setForeground(QtGui.QBrush(QtGui.QColor('blue')))
                table.setItem(i, j, item)

        if num != 3:
            self.updateMerge()
        else:
            self.setStatusTip("Updated Preview")
    
    def updateMerge(self):
        self.mm.updateData()
        self.updateTable(self.table3, self.mm, 3)

def main():
    app = QtWidgets.QApplication(sys.argv)
    window = MainWindow()
    window.show()
    app.exec_()

if __name__ == '__main__':
    main()
